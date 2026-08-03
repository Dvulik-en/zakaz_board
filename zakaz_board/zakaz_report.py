"""
zakaz_report.py
================
Полный цикл: план ПДО (месяц) -> список заказов/изделий из плана ->
для каждого спрашиваем папку с xlsm-спецификациями -> собираем раскрой-отчёт
с датой (из плана), полем УП (программа раскроя, заполняете сами при раскладке)
и статусом (заполняет сортировщик отдельно). При повторном запуске уже
введённые УП и Статус НЕ теряются — переносятся в новый отчёт по коду детали.

ШАГ 1. Разбор формата размеров и материала стал умнее:
  - "300х452"        -> прямоугольная заготовка
  - "Ø63хØ26"         -> кольцо/диск: первое число — наружный диаметр,
                         второе — внутренний (площадь считается по наружному)
  - "Ø140"            -> сплошной диск
  - "Лист г/к 10мм,09Г2С"          -> обычный лист, стандартный запас (6000х1500)
  - "Лист г/к 16х2000х6000мм,09Г2С" -> НЕСТАНДАРТНЫЙ лист 2000х6000мм,
    толщина 16мм — выделяется в отдельную группу, не мешается со стандартным
    металлом той же толщины (это другая заготовка на складе)

Использование:
    python zakaz_report.py "Заготовительный_участок.xlsx"

Дальше скрипт покажет список заказов/изделий из плана (лист "Плазма ( план )"),
для каждого спросит путь к папке с xlsm (Enter — пропустить), соберёт отчёт
report_<дата запуска>.xlsx рядом со скриптом.

Зависимости: openpyxl (pip install openpyxl)
"""

import sys
import os
import re
import glob
import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Нужен openpyxl: pip install openpyxl")
    sys.exit(1)

SHEET_NAME = "Лист1"                 # лист внутри xlsm-спецификации изделия
PLAN_SHEET_NAME = "Плазма ( план )"  # лист плана ПДО в файле "Заготовительный участок"
STD_SHEET_WIDTH_MM = 6000
STD_SHEET_HEIGHT_MM = 1500
STD_SHEET_AREA_M2 = (STD_SHEET_WIDTH_MM / 1000) * (STD_SHEET_HEIGHT_MM / 1000)

REQUIRED_HEADERS = [
    "КД продукции", "Наименование продукции", "Материал",
    "Размеры заготовки,мм", "Количество в узле,шт", "операция 005",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PRODUCT_FILL = PatternFill("solid", fgColor="D9E1F2")
PRODUCT_FONT = Font(bold=True)
SUBTOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
SUBTOTAL_FONT = Font(bold=True, italic=True)
DONE_FILL = PatternFill("solid", fgColor="C6EFCE")
NONSTD_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN_BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)

DETAIL_COLS = ["Статус", "УП", "Код детали", "Наименование", "Размер заготовки", "Кол-во", "Площадь,м2"]
COL_STATUS, COL_UP, COL_CODE, COL_NAME, COL_SIZE, COL_QTY, COL_AREA = range(1, 8)


# ---------------------------------------------------------------- парсинг ---

def parse_material(mat):
    """Возвращает (толщина_мм, марка, нестанд_лист_или_None)."""
    if not mat:
        return None, None, None
    mat = str(mat)
    m3 = re.search(r"(\d+(?:[.,]\d+)?)[xх](\d+(?:[.,]\d+)?)[xх](\d+(?:[.,]\d+)?)\s*мм", mat)
    if m3:
        thickness = float(m3.group(1).replace(",", "."))
        custom_sheet = (float(m3.group(2).replace(",", ".")), float(m3.group(3).replace(",", ".")))
    else:
        m1 = re.search(r"(\d+(?:[.,]\d+)?)\s*мм", mat)
        thickness = float(m1.group(1).replace(",", ".")) if m1 else None
        custom_sheet = None
    gm = re.search(r",\s*([А-Яа-яA-Za-z0-9]+)\s*$", mat)
    grade = gm.group(1) if gm else None
    return thickness, grade, custom_sheet


def parse_size(size):
    """Возвращает словарь с типом заготовки: rect / circle / unknown."""
    if not size:
        return {"type": None}
    s = str(size).replace(" ", "")
    diams = re.findall(r"[ØøD](\d+(?:[.,]\d+)?)", s)
    if diams:
        vals = [float(x.replace(",", ".")) for x in diams]
        if len(vals) >= 2:
            return {"type": "circle", "d_outer": max(vals), "d_inner": min(vals)}
        return {"type": "circle", "d_outer": vals[0], "d_inner": 0}
    m = re.match(r"^(\d+[.,]?\d*)[xх](\d+[.,]?\d*)$", s)
    if m:
        return {"type": "rect", "a": float(m.group(1).replace(",", ".")), "b": float(m.group(2).replace(",", "."))}
    return {"type": "unknown", "raw": s}


def size_to_str(size_info):
    if size_info["type"] == "rect":
        return f'{size_info["a"]:g}x{size_info["b"]:g}'
    if size_info["type"] == "circle":
        if size_info["d_inner"]:
            return f'Ø{size_info["d_outer"]:g}хØ{size_info["d_inner"]:g}'
        return f'Ø{size_info["d_outer"]:g}'
    if size_info["type"] == "unknown":
        return size_info.get("raw", "?")
    return "?"


def area_m2(size_info, qty):
    import math
    if size_info["type"] == "rect":
        return size_info["a"] * size_info["b"] * qty / 1_000_000
    if size_info["type"] == "circle":
        return math.pi / 4 * (size_info["d_outer"] ** 2) * qty / 1_000_000
    return 0.0


# ------------------------------------------------------- разбор xlsm изделий ---

def extract_plasma_parts(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"    ! Не удалось открыть {os.path.basename(filepath)}: {e}")
        return []
    if SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[SHEET_NAME]

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            headers[str(v).strip()] = c
    if any(h not in headers for h in REQUIRED_HEADERS):
        return []

    col_code, col_name = headers["КД продукции"], headers["Наименование продукции"]
    col_mat, col_size = headers["Материал"], headers["Размеры заготовки,мм"]
    col_qty, col_op = headers["Количество в узле,шт"], headers["операция 005"]

    rows = []
    for r in range(2, ws.max_row + 1):
        op = ws.cell(row=r, column=col_op).value
        if op and "плазм" in str(op).lower():
            qty = ws.cell(row=r, column=col_qty).value or 0
            material_raw = ws.cell(row=r, column=col_mat).value
            thickness, grade, custom_sheet = parse_material(material_raw)
            size_info = parse_size(ws.cell(row=r, column=col_size).value)
            rows.append({
                "code": ws.cell(row=r, column=col_code).value,
                "name": ws.cell(row=r, column=col_name).value,
                "grade": grade, "thickness": thickness, "custom_sheet": custom_sheet,
                "material_raw": (material_raw or "").strip(),
                "size_info": size_info, "qty": qty,
                "area_m2": area_m2(size_info, qty),
            })
    return rows


def find_produkciya_folder(order_folder):
    for root, dirs, _ in os.walk(order_folder):
        for d in dirs:
            if d.strip().lower() == "продукция":
                return os.path.join(root, d)
    return None


def _normalize_text(s):
    s = str(s or "").lower()
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def match_product_folders(target_produkciya, product_dirs):
    """Возвращает список (папка, оценка_совпадения 0..1), отсортированный по убыванию."""
    target = _normalize_text(target_produkciya)
    target_tokens = set(target.split())
    scored = []
    for d in product_dirs:
        dn = _normalize_text(d)
        if not dn:
            continue
        if target and (target in dn or dn in target):
            scored.append((d, 1.0))
            continue
        dn_tokens = set(dn.split())
        if not dn_tokens or not target_tokens:
            scored.append((d, 0.0))
            continue
        overlap = len(target_tokens & dn_tokens) / len(target_tokens | dn_tokens)
        scored.append((d, overlap))
    scored.sort(key=lambda x: -x[1])
    return scored


def choose_product_dirs(product_dirs, target_produkciya):
    """
    Если задана конкретная продукция из плана — пытается найти именно её подпапку(и),
    а не забирать всё, что лежит в ПРОДУКЦИЯ. При неоднозначности — спрашивает у человека.
    """
    if not target_produkciya or not product_dirs:
        return product_dirs

    scored = match_product_folders(target_produkciya, product_dirs)
    exact = [d for d, s in scored if s >= 0.99]
    if len(exact) == 1:
        return exact

    good = [d for d, s in scored if s >= 0.4]
    if len(good) == 1:
        print(f"    -> по названию нашлась папка: «{good[0]}»")
        return good

    print(f"\n    Для «{target_produkciya}» не нашлось однозначного совпадения папки.")
    print("    Варианты в этой папке ПРОДУКЦИЯ:")
    for i, d in enumerate(product_dirs, start=1):
        print(f"      {i}. {d}")
    choice = input("    Введите номер(а) через запятую (нужную папку/папки), Enter — взять всё: ").strip()
    if not choice:
        return product_dirs
    try:
        idxs = [int(x.strip()) for x in choice.split(",") if x.strip()]
        return [product_dirs[i - 1] for i in idxs if 1 <= i <= len(product_dirs)]
    except (ValueError, IndexError):
        print("    Не разобрал ввод, беру всё из папки.")
        return product_dirs


def build_product_groups(folder, target_produkciya=None):
    """
    Заказ (папка) -> {изделие: {(grade,thickness,custom_sheet): {parts,qty,area_m2}}}
    Логика поиска папок такая же, как раньше: ПРОДУКЦИЯ -> подпапка изделия -> xlsm.
    Если передан target_produkciya (конкретное изделие из плана ПДО) — берутся только
    подпапки, соответствующие именно этому изделию, а не вся ПРОДУКЦИЯ целиком.
    """
    structure = {}
    product_root = find_produkciya_folder(folder) or folder
    entries = sorted(os.listdir(product_root))
    product_dirs = [e for e in entries if os.path.isdir(os.path.join(product_root, e))]
    loose_files = [e for e in entries if e.lower().endswith(".xlsm") and not e.startswith("~$")]

    product_dirs = choose_product_dirs(product_dirs, target_produkciya)

    def add(bucket, rows):
        for row in rows:
            key = (row["grade"], row["thickness"], row["custom_sheet"])
            g = bucket.setdefault(key, {"parts": [], "qty": 0, "area_m2": 0.0})
            g["parts"].append(row)
            g["qty"] += row["qty"]
            g["area_m2"] += row["area_m2"]

    for pdir in product_dirs:
        full_path = os.path.join(product_root, pdir)
        files = [f for f in glob.glob(os.path.join(full_path, "**", "*.xlsm"), recursive=True)
                 if not os.path.basename(f).startswith("~$")]
        if not files:
            continue
        bucket = structure.setdefault(pdir, {})
        for f in sorted(files):
            add(bucket, extract_plasma_parts(f))

    for lf in loose_files:
        rows = extract_plasma_parts(os.path.join(product_root, lf))
        if rows:
            name = os.path.splitext(lf)[0]
            bucket = structure.setdefault(name, {})
            add(bucket, rows)

    return structure


# -------------------------------------------------------------- план ПДО ---

def parse_plan(ws):
    """Возвращает список {date, zakazchik, zavno, produkciya, kolvo}."""
    entries = []
    current_date = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and re.match(r"^\d{2}\.\d{2}\.\d{4}", a.strip()):
            current_date = a.strip().split()[0]
            continue
        if a == "№":
            continue
        produkciya = ws.cell(row=r, column=4).value
        if produkciya:
            entries.append({
                "date": current_date,
                "zakazchik": ws.cell(row=r, column=2).value,
                "zavno": ws.cell(row=r, column=3).value,
                "produkciya": str(produkciya).strip(),
                "kolvo": ws.cell(row=r, column=5).value,
            })
    return entries


def group_plan_entries(entries):
    """label (заказчик-решение + продукция) -> {dates:set, zavno:set, kolvo_total}"""
    groups = defaultdict(lambda: {"dates": set(), "zavno": set(), "entries": []})
    for e in entries:
        label = f'{e["zakazchik"]} — {e["produkciya"]}'
        g = groups[label]
        if e["date"]:
            g["dates"].add(e["date"])
        if e["zavno"]:
            g["zavno"].add(str(e["zavno"]))
        g["entries"].append(e)
    return groups


# ------------------------------------------------------- сохранение существующих отметок ---

def load_previous_marks(path):
    """Читает уже сохранённый отчёт (если есть) и возвращает {(изделие,код): (статус,УП)}."""
    marks = {}
    if not os.path.exists(path):
        return marks
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return marks
    for sheet_name in wb.sheetnames:
        if sheet_name == "Сводка":
            continue
        ws = wb[sheet_name]
        current_product = None
        for r in range(1, ws.max_row + 1):
            v1 = ws.cell(row=r, column=1).value
            if isinstance(v1, str) and "ИЗДЕЛИЕ:" in v1:
                name_part = v1.split("ИЗДЕЛИЕ:", 1)[1]
                name_part = name_part.split(" | дата", 1)[0].strip()
                current_product = f"ИЗДЕЛИЕ: {name_part}"
                continue
            if v1 in ("Не резано", "В раскрое", "Вырезано"):
                code = ws.cell(row=r, column=COL_CODE).value
                up = ws.cell(row=r, column=COL_UP).value
                if code and current_product:
                    marks[(current_product, str(code))] = (v1, up)
    return marks


# ------------------------------------------------------------------- вывод ---

def sanitize_sheet_name(name, used):
    bad = '\\/?*[]:'
    clean = "".join(c for c in name if c not in bad).strip()[:31] or "Лист"
    base, i = clean, 2
    while clean in used:
        suf = f"_{i}"
        clean = base[: 31 - len(suf)] + suf
        i += 1
    used.add(clean)
    return clean


def write_group_sheet(wb, key, entries, dates_by_product, used_names):
    grade, thickness, custom_sheet = key
    th_str = f"{thickness:g}" if thickness is not None else "?"
    title = f"{grade or '?'} {th_str}мм"
    if custom_sheet:
        title += f" НЕСТАНД {custom_sheet[0]:g}x{custom_sheet[1]:g}"
    ws = wb.create_sheet(sanitize_sheet_name(title, used_names))

    total_qty = sum(g["qty"] for _, _, g in entries)
    total_area = sum(g["area_m2"] for _, _, g in entries)
    sheet_area = (custom_sheet[0] / 1000 * custom_sheet[1] / 1000) if custom_sheet else STD_SHEET_AREA_M2
    ncols = len(DETAIL_COLS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1)
    tc.value = (f"{title} — {total_qty} дет., {total_area:.2f} м² "
                f"(~{total_area / sheet_area:.2f} листа)")
    tc.font = Font(bold=True, size=13)
    if custom_sheet:
        tc.fill = NONSTD_FILL

    row = 3
    for order_name, product_name, g in entries:
        dates = dates_by_product.get(product_name)
        date_str = f" | дата по плану: {', '.join(sorted(dates))}" if dates else ""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1)
        c.value = f"ЗАКАЗ: {order_name}   /   ИЗДЕЛИЕ: {product_name}{date_str}"
        c.fill = PRODUCT_FILL
        c.font = PRODUCT_FONT
        row += 1

        for ci, col_name in enumerate(DETAIL_COLS, start=1):
            hc = ws.cell(row=row, column=ci, value=col_name)
            hc.fill = HEADER_FILL
            hc.font = HEADER_FONT
            hc.border = THIN_BORDER
        row += 1
        first_data_row = row

        for p in sorted(g["parts"], key=lambda x: str(x["code"])):
            status, up = g.get("marks", {}).get(str(p["code"]), ("", ""))
            values = [status, up, p["code"], p["name"], size_to_str(p["size_info"]),
                      p["qty"], round(p["area_m2"], 3)]
            for ci, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.border = THIN_BORDER
            row += 1
        last_row = row - 1

        if last_row >= first_data_row:
            dv_status = DataValidation(type="list", formula1='"Не резано,В раскрое,Вырезано"', allow_blank=True)
            ws.add_data_validation(dv_status)
            rng = f"A{first_data_row}:A{last_row}"
            dv_status.add(rng)
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Вырезано"'], fill=DONE_FILL))

            dv_up = DataValidation(type="textLength", operator="lessThanOrEqual", formula1="8",
                                    allow_blank=True, showErrorMessage=True,
                                    errorTitle="Слишком длинно", error="УП — максимум 8 символов")
            ws.add_data_validation(dv_up)
            dv_up.add(f"B{first_data_row}:B{last_row}")

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols - 2)
        sc = ws.cell(row=row, column=1)
        sc.value = f"Итого по изделию: {g['qty']} дет."
        sc.fill, sc.font = SUBTOTAL_FILL, SUBTOTAL_FONT
        ac = ws.cell(row=row, column=ncols)
        ac.value = round(g["area_m2"], 3)
        ac.fill, ac.font = SUBTOTAL_FILL, SUBTOTAL_FONT
        row += 2

    widths = [12, 10, 26, 24, 16, 8, 12]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"


def save_report(orders_data, dates_by_product, out_path):
    previous = load_previous_marks(out_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_sum = wb.create_sheet("Сводка")
    ws_sum.append(["Заказ", "Изделие", "Марка", "Толщина,мм", "Нестанд.лист", "Деталей", "Площадь,м2"])
    for c in range(1, 8):
        cell = ws_sum.cell(row=1, column=c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT

    by_key = defaultdict(list)  # key -> [(order,product,group_with_marks)]
    for order_name, structure in orders_data.items():
        for product_name, groups in structure.items():
            for key, g in groups.items():
                gg = dict(g)
                gg["marks"] = {str(p["code"]): previous.get((f"ИЗДЕЛИЕ: {product_name}", str(p["code"])),
                                                              ("", "")) for p in g["parts"]}
                by_key[key].append((order_name, product_name, gg))
                grade, thickness, custom_sheet = key
                ws_sum.append([order_name, product_name, grade, thickness,
                               f"{custom_sheet[0]:g}x{custom_sheet[1]:g}" if custom_sheet else "",
                               gg["qty"], round(gg["area_m2"], 3)])

    for ci, w in enumerate([20, 40, 10, 12, 14, 10, 12], start=1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.freeze_panes = "A2"

    used_names = {"Сводка"}
    for key in sorted(by_key.keys(), key=lambda k: (k[0] or "", k[1] or 0)):
        write_group_sheet(wb, key, by_key[key], dates_by_product, used_names)

    wb.save(out_path)
    carried = sum(1 for v in previous.values() if v[0] or v[1])
    print(f"\nСохранено: {out_path}")
    if carried:
        print(f"Перенесено из предыдущего отчёта отметок Статус/УП: {carried}")


# -------------------------------------------------------------------- main ---

def main():
    if len(sys.argv) < 2:
        print('Использование: python zakaz_report.py "Заготовительный_участок.xlsx"')
        sys.exit(1)

    plan_path = sys.argv[1]
    wb = openpyxl.load_workbook(plan_path, data_only=True)
    if PLAN_SHEET_NAME not in wb.sheetnames:
        print(f"В файле нет листа '{PLAN_SHEET_NAME}'")
        sys.exit(1)

    entries = parse_plan(wb[PLAN_SHEET_NAME])
    groups = group_plan_entries(entries)

    print(f"В плане найдено {len(groups)} заказов/изделий:\n")
    labels = sorted(groups.keys())
    for i, label in enumerate(labels, start=1):
        dates = ", ".join(sorted(groups[label]["dates"]))
        print(f"  {i}. {label}   [{dates}]")

    orders_data = {}          # order_name -> {product_name: {...}}
    dates_by_product = {}     # "ИЗДЕЛИЕ: <имя папки изделия>" -> set(dates)  (заполнится ниже)

    print("\nДля каждого пункта укажите путь к папке заказа (Enter — пропустить).")
    for label in labels:
        path = input(f"\n[{label}]\nпапка: ").strip().strip('"')
        if not path:
            continue
        if not os.path.isdir(path):
            print("  ! Такой папки нет, пропускаю")
            continue
        order_name = os.path.basename(os.path.normpath(path))
        target_produkciya = groups[label]["entries"][0]["produkciya"] if groups[label]["entries"] else None
        structure = build_product_groups(path, target_produkciya=target_produkciya)
        if not structure:
            print("  ! В папке не нашлось изделий с деталями на плазму")
            continue
        # одна и та же папка заказа может использоваться повторно для другого изделия
        # из той же папки ПРОДУКЦИЯ — объединяем, а не перетираем предыдущий результат
        orders_data.setdefault(order_name, {}).update(structure)
        for product_name in structure:
            dates_by_product[f"ИЗДЕЛИЕ: {product_name}"] = groups[label]["dates"]
        print(f"  ok: найдено изделий — {len(structure)}")

    if not orders_data:
        print("\nНичего не обработано, выхожу.")
        return

    out_path = os.path.join(os.getcwd(), "report.xlsx")
    save_report(orders_data, dates_by_product, out_path)


if __name__ == "__main__":
    main()
